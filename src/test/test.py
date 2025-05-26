import pandas as pd
from fuzzywuzzy import process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# Читаем файл Excel
df = pd.read_excel('для_поиска_повторных_артикулов_2.xlsx')

# Находим записи с повторяющимися артикулами
duplicates = df[df.duplicated(subset='Артикул', keep=False)]

# Записываем записи с повторяющимися артикулами в новый файл Excel
duplicates.to_excel('повторные_артикулы.xlsx', index=False)



# df1 = pd.read_excel('Артикула_проверить.xlsx') # Файл для проверки
# df2 = pd.read_excel('test.xlsx') # Файл с верными артикулами
#
# # df2 = df2.drop_duplicates()
# # df2.to_excel('test.xlsx', index=False)
#
# results = []
#
# for index, row in df1.iterrows():
#     name = row['Наименование']
#     best_match = process.extractOne(name, df2['Наименование'])
#
#     if best_match[1] >= 80:  # Установите порог совпадения (например, 80%)
#         correct_article = df2.loc[df2['Наименование'] == best_match[0], 'Артикул'].values[0]
#         results.append((name, row['Артикул'], correct_article, best_match[0]))
#         print(f"Наименование: {name}, Артикул (из 1-го файла): {row['Артикул']}, Верный артикул: {correct_article}, Наименование (из 2-го файла): {best_match[0]}")
#     else:
#         results.append((name, row['Артикул'], None))  # Если совпадение не найдено
#
#
# result_df = pd.DataFrame(results, columns=['Наименование', 'Артикул (из 1-го файла)', 'Верный артикул', 'Наименование (из 2-го файла)'])
#
# result_df.to_excel('результаты.xlsx', index=False)

# Загрузка данных из Excel файла
# file_path = 'результаты.xlsx'
# df = pd.read_excel(file_path)
#
# # Загрузка книги для редактирования
# wb = load_workbook(file_path)
# ws = wb.active
#
# # Определение цвета заливки для выделения строк
# highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Желтый цвет
#
# # Проход по строкам DataFrame и выделение цветом несоответствующих строк
# for index, row in df.iterrows():
#     if row['Артикул (из 1-го файла)'] != row['Верный артикул']:
#         # Выделение всей строки в Excel
#         for cell in ws[index + 2]:  # +2, так как индексация в DataFrame начинается с 0, а в Excel с 1 и есть заголовок
#             cell.fill = highlight_fill
#
# # Сохранение изменений в новом файле
# wb.save('результаты_с_выделением.xlsx')